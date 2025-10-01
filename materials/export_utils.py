"""
Утилиты для экспорта распознанных OCR данных в табличные форматы (Excel, CSV)
"""

import os
import csv
import json
import logging
from typing import Dict, List, Any, Optional
from io import BytesIO, StringIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)

# Попытка импорта библиотек для работы с Excel
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_EXCEL_SUPPORT = True
    logger.info("Excel библиотеки успешно загружены")
except ImportError as e:
    HAS_EXCEL_SUPPORT = False
    logger.warning(f"Excel библиотеки не установлены: {e}, экспорт в Excel недоступен")


class TTNExportService:
    """
    Сервис для экспорта распознанных данных ТТН в различные форматы
    """
    
    def __init__(self):
        # Маппинг полей для экспорта
        self.field_mapping = {
            'document_number': 'Номер ТТН',
            'document_date': 'Дата ТТН',
            'sender_name': 'Отправитель',
            'sender_inn': 'ИНН отправителя',
            'receiver_name': 'Получатель',
            'receiver_inn': 'ИНН получателя',
            'vehicle_number': 'Номер ТС',
            'driver_name': 'ФИО водителя',
            'cargo_description': 'Описание груза',
            'cargo_weight': 'Вес груза (кг)',
            'cargo_volume': 'Объем груза (м³)',
            'packages_count': 'Количество мест',
            'processing_status': 'Статус обработки',
            'ocr_confidence': 'Уверенность OCR (%)',
            'manual_verification_required': 'Требует проверки',
            'created_at': 'Дата создания',
            'processed_at': 'Дата обработки'
        }
        
    def export_to_csv(self, queryset, include_ocr_details: bool = False) -> HttpResponse:
        """
        Экспорт данных в CSV формат
        
        Args:
            queryset: QuerySet с объектами для экспорта
            include_ocr_details: Включать ли детали OCR обработки
            
        Returns:
            HttpResponse с CSV файлом
        """
        logger.info(f"Начало экспорта {queryset.count()} записей в CSV")
        
        # Создаем HTTP ответ с CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="ttn_export_{timestamp}.csv"'
        
        # Добавляем BOM для корректного отображения кириллицы в Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Заголовки
        headers = list(self.field_mapping.values())
        if include_ocr_details:
            headers.extend(['Извлеченные поля OCR', 'Уверенность по полям', 'Ошибки валидации'])
        
        writer.writerow(headers)
        
        # Данные
        for obj in queryset:
            row_data = self._prepare_row_data(obj, include_ocr_details)
            writer.writerow(row_data)
        
        logger.info(f"CSV экспорт завершен: {queryset.count()} записей")
        return response
    
    def export_to_excel(self, queryset, include_ocr_details: bool = False) -> HttpResponse:
        """
        Экспорт данных в Excel формат
        
        Args:
            queryset: QuerySet с объектами для экспорта
            include_ocr_details: Включать ли детали OCR обработки
            
        Returns:
            HttpResponse с Excel файлом
        """
        if not HAS_EXCEL_SUPPORT:
            raise ValueError("Excel библиотеки не установлены")
        
        logger.info(f"Начало экспорта {queryset.count()} записей в Excel")
        
        # Подготавливаем данные
        data = []
        for obj in queryset:
            row_data = self._prepare_dict_data(obj, include_ocr_details)
            data.append(row_data)
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Переименовываем колонки
        column_mapping = self.field_mapping.copy()
        if include_ocr_details:
            column_mapping.update({
                'extracted_fields': 'Извлеченные поля OCR',
                'field_confidences': 'Уверенность по полям',
                'validation_errors': 'Ошибки валидации'
            })
        
        # Применяем маппинг только для существующих колонок
        existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
        df = df.rename(columns=existing_columns)
        
        # Создаем Excel файл в памяти
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='ТТН_Данные', index=False)
            
            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets['ТТН_Данные']
            
            # Форматирование заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:  # Первая строка (заголовки)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Автоматическая ширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Подготавливаем HTTP ответ
        output.seek(0)
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ttn_export_{timestamp}.xlsx"'
        
        logger.info(f"Excel экспорт завершен: {queryset.count()} записей")
        return response
    
    def export_summary_to_excel(self, queryset) -> HttpResponse:
        """
        Экспорт сводной информации по обработке ТТН в Excel
        
        Args:
            queryset: QuerySet с объектами TransportDocument
            
        Returns:
            HttpResponse с Excel файлом со сводкой
        """
        if not HAS_EXCEL_SUPPORT:
            raise ValueError("Excel библиотеки не установлены")
        
        logger.info(f"Создание сводного отчета по {queryset.count()} ТТН")
        
        # Подготавливаем данные для сводки
        summary_data = self._prepare_summary_data(queryset)
        
        output = BytesIO()
        workbook = openpyxl.Workbook()
        
        # Лист со сводкой
        summary_sheet = workbook.active
        summary_sheet.title = "Сводка"
        
        # Заголовок отчета
        summary_sheet['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ОБРАБОТКЕ ТТН"
        summary_sheet['A1'].font = Font(size=16, bold=True)
        summary_sheet.merge_cells('A1:C1')
        
        summary_sheet['A2'] = f"Дата формирования: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
        summary_sheet.merge_cells('A2:C2')
        
        # Общая статистика
        row = 4
        summary_sheet[f'A{row}'] = "ОБЩАЯ СТАТИСТИКА"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        
        for key, value in summary_data['general_stats'].items():
            row += 1
            summary_sheet[f'A{row}'] = key
            summary_sheet[f'B{row}'] = value
        
        # Статистика по статусам
        row += 2
        summary_sheet[f'A{row}'] = "СТАТИСТИКА ПО СТАТУСАМ ОБРАБОТКИ"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        
        for status, count in summary_data['status_stats'].items():
            row += 1
            summary_sheet[f'A{row}'] = status
            summary_sheet[f'B{row}'] = count
        
        # Статистика по уверенности OCR
        row += 2
        summary_sheet[f'A{row}'] = "СТАТИСТИКА ПО КАЧЕСТВУ РАСПОЗНАВАНИЯ"
        summary_sheet[f'A{row}'].font = Font(bold=True)
        
        for range_name, count in summary_data['confidence_stats'].items():
            row += 1
            summary_sheet[f'A{row}'] = range_name
            summary_sheet[f'B{row}'] = count
        
        # Детальные данные на отдельном листе
        details_sheet = workbook.create_sheet("Детальные данные")
        
        # Подготавливаем детальные данные
        detail_data = []
        for obj in queryset:
            detail_data.append(self._prepare_dict_data(obj, include_ocr_details=True))
        
        if detail_data:
            df = pd.DataFrame(detail_data)
            df = df.rename(columns=self.field_mapping)
            
            # Записываем данные в лист
            for r in dataframe_to_rows(df, index=False, header=True):
                details_sheet.append(r)
            
            # Форматирование заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in details_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Сохраняем в BytesIO
        workbook.save(output)
        output.seek(0)
        
        # HTTP ответ
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ttn_summary_{timestamp}.xlsx"'
        
        logger.info(f"Сводный отчет создан для {queryset.count()} записей")
        return response
    
    def _prepare_row_data(self, obj, include_ocr_details: bool = False) -> List[Any]:
        """Подготовить данные строки для CSV экспорта"""
        from .models import TransportDocument
        
        if isinstance(obj, TransportDocument):
            transport_doc = obj
        else:
            transport_doc = obj.transport_document
        
        row_data = [
            transport_doc.document_number,
            transport_doc.document_date.strftime('%d.%m.%Y') if transport_doc.document_date else '',
            transport_doc.sender_name,
            transport_doc.sender_inn,
            transport_doc.receiver_name,
            transport_doc.receiver_inn,
            transport_doc.vehicle_number,
            transport_doc.driver_name,
            transport_doc.cargo_description,
            transport_doc.cargo_weight,
            transport_doc.cargo_volume,
            transport_doc.packages_count,
            transport_doc.get_processing_status_display(),
            f"{transport_doc.ocr_confidence:.2f}" if transport_doc.ocr_confidence else '',
            "Да" if transport_doc.manual_verification_required else "Нет",
            transport_doc.created_at.strftime('%d.%m.%Y %H:%M') if transport_doc.created_at else '',
            transport_doc.updated_at.strftime('%d.%m.%Y %H:%M') if transport_doc.updated_at else '',
        ]
        
        if include_ocr_details and hasattr(transport_doc, 'photos'):
            # Собираем OCR данные со всех фотографий
            ocr_fields = {}
            confidence_data = {}
            validation_errors = []
            
            for photo in transport_doc.photos.all():
                if hasattr(photo, 'ocr_result'):
                    ocr_result = photo.ocr_result
                    ocr_fields.update(ocr_result.extracted_fields)
                    confidence_data.update(ocr_result.field_confidences)
                    validation_errors.extend(ocr_result.validation_errors)
            
            row_data.extend([
                json.dumps(ocr_fields, ensure_ascii=False),
                json.dumps(confidence_data, ensure_ascii=False),
                '; '.join(validation_errors) if validation_errors else ''
            ])
        
        return row_data
    
    def _prepare_dict_data(self, obj, include_ocr_details: bool = False) -> Dict[str, Any]:
        """Подготовить данные в виде словаря для DataFrame"""
        from .models import TransportDocument
        
        if isinstance(obj, TransportDocument):
            transport_doc = obj
        else:
            transport_doc = obj.transport_document
        
        data = {
            'document_number': transport_doc.document_number,
            'document_date': transport_doc.document_date.strftime('%d.%m.%Y') if transport_doc.document_date else '',
            'sender_name': transport_doc.sender_name,
            'sender_inn': transport_doc.sender_inn,
            'receiver_name': transport_doc.receiver_name,
            'receiver_inn': transport_doc.receiver_inn,
            'vehicle_number': transport_doc.vehicle_number,
            'driver_name': transport_doc.driver_name,
            'cargo_description': transport_doc.cargo_description,
            'cargo_weight': transport_doc.cargo_weight,
            'cargo_volume': transport_doc.cargo_volume,
            'packages_count': transport_doc.packages_count,
            'processing_status': transport_doc.get_processing_status_display(),
            'ocr_confidence': f"{transport_doc.ocr_confidence:.2f}" if transport_doc.ocr_confidence else '',
            'manual_verification_required': "Да" if transport_doc.manual_verification_required else "Нет",
            'created_at': transport_doc.created_at.strftime('%d.%m.%Y %H:%M') if transport_doc.created_at else '',
            'processed_at': transport_doc.updated_at.strftime('%d.%m.%Y %H:%M') if transport_doc.updated_at else '',
        }
        
        if include_ocr_details and hasattr(transport_doc, 'photos'):
            # Собираем OCR данные
            ocr_fields = {}
            confidence_data = {}
            validation_errors = []
            
            for photo in transport_doc.photos.all():
                if hasattr(photo, 'ocr_result'):
                    ocr_result = photo.ocr_result
                    ocr_fields.update(ocr_result.extracted_fields)
                    confidence_data.update(ocr_result.field_confidences)
                    validation_errors.extend(ocr_result.validation_errors)
            
            data.update({
                'extracted_fields': json.dumps(ocr_fields, ensure_ascii=False),
                'field_confidences': json.dumps(confidence_data, ensure_ascii=False),
                'validation_errors': '; '.join(validation_errors) if validation_errors else ''
            })
        
        return data
    
    def _prepare_summary_data(self, queryset) -> Dict[str, Any]:
        """Подготовить сводные данные для отчета"""
        total_count = queryset.count()
        
        # Общая статистика
        general_stats = {
            'Всего ТТН': total_count,
            'Обработано': queryset.filter(processing_status='processed').count(),
            'Проверено': queryset.filter(processing_status='verified').count(),
            'Требует ручной проверки': queryset.filter(manual_verification_required=True).count(),
            'С ошибками': queryset.filter(processing_status='error').count(),
        }
        
        # Статистика по статусам
        status_stats = {}
        for status_code, status_name in queryset.model._meta.get_field('processing_status').choices:
            count = queryset.filter(processing_status=status_code).count()
            if count > 0:
                status_stats[status_name] = count
        
        # Статистика по уверенности OCR
        confidence_stats = {
            'Высокая уверенность (>80%)': queryset.filter(ocr_confidence__gt=80).count(),
            'Средняя уверенность (60-80%)': queryset.filter(ocr_confidence__gt=60, ocr_confidence__lte=80).count(),
            'Низкая уверенность (40-60%)': queryset.filter(ocr_confidence__gt=40, ocr_confidence__lte=60).count(),
            'Очень низкая уверенность (<40%)': queryset.filter(ocr_confidence__lte=40).count(),
            'Не определена': queryset.filter(ocr_confidence__isnull=True).count(),
        }
        
        return {
            'general_stats': general_stats,
            'status_stats': status_stats,
            'confidence_stats': confidence_stats
        }


# Глобальный экземпляр сервиса экспорта
ttn_export_service = TTNExportService()